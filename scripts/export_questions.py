from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(
    "/Users/gaoxing/Desktop/3-2-10重点人员、一线人员--生产变电安规通用+通信安规+规章制度+严重违章释义+生产变电典型作业场景保人身措施.xlsx"
)
OUTPUT = Path(__file__).resolve().parents[1] / "data" / "questions.js"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_options(raw: str) -> list[dict[str, str]]:
    if not raw:
        return []

    options = []
    for part in raw.split("|"):
        item = part.strip()
        if not item:
            continue
        if "-" in item:
            key, text = item.split("-", 1)
        elif "．" in item:
            key, text = item.split("．", 1)
        elif "." in item:
            key, text = item.split(".", 1)
        else:
            key, text = "", item
        options.append({"key": key.strip().upper(), "text": text.strip()})
    return options


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["题库格式"]
    headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: pos for pos, name in enumerate(headers)}

    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stem = clean(row[index["题干"]])
        if not stem:
            continue

        answer = clean(row[index["答案"]]).upper().replace("，", ",")
        question = {
            "id": int(row[index["序号"]]) if row[index["序号"]] else len(questions) + 1,
            "outline1": clean(row[index["一级纲要"]]),
            "outline2": clean(row[index["二级纲要"]]),
            "category": clean(row[index["题目分类"]]) or "未分类",
            "type": clean(row[index["题型"]]) or "未知题型",
            "stem": stem,
            "options": parse_options(clean(row[index["选项"]])),
            "answer": answer,
            "basis": clean(row[index["题目依据"]]),
            "score": clean(row[index["试题分数"]]) or "1",
            "code": clean(row[index["试题编码"]]),
            "note": clean(row[index["备注"]]),
        }
        questions.append(question)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(questions, ensure_ascii=False, separators=(",", ":"))
    OUTPUT.write_text(f"window.QUESTION_BANK = {payload};\n", encoding="utf-8")
    print(f"Exported {len(questions)} questions to {OUTPUT}")


if __name__ == "__main__":
    main()
